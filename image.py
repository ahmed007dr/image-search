import os
import openpyxl
from google_images_download import google_images_download
import time

def create_output_directory(output_directory):
    # Create output directory if it doesn't exist
    if not os.path.exists(output_directory):
        os.makedirs(output_directory)

def load_excel_sheet(excel_file):
    # Load Excel file
    return openpyxl.load_workbook(excel_file).active

def search_and_download_images(sheet, output_directory):
    # Initialize GoogleImagesDownload
    response = google_images_download.googleimagesdownload()

    # Loop through each row in the first column of the Excel sheet
    for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row, values_only=True):
        name = row[0]
        if name:
            print(f"Searching for images of '{name}'...")
            # Search for images on Google Images
            arguments = {"keywords": name, "limit": 1, "output_directory": output_directory}
            try:
                paths = response.download(arguments)
                print(f"Downloaded image for '{name}'")
            except Exception as e:
                print(f"Error downloading image for '{name}': {e}")
            
            # Introduce a delay between each image download request
            time.sleep(3)  # Adjust the delay time as needed

if __name__ == "__main__":
    # Specify the path to your Excel file
    excel_file = "F:/django-5/image/src/Drug.xlsx"
    
    # Specify the output directory for downloaded images
    output_directory = "F:/django-5/image/src/output_directory"

    create_output_directory(output_directory)
    sheet = load_excel_sheet(excel_file)
    search_and_download_images(sheet, output_directory)
